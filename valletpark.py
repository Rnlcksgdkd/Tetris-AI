
import win32com.client as win32
from openpyxl import load_workbook



### 주피터노트북에 win32 설치가 안되는 현상 ??

# xls > xlsx 변환
def convert2xlsx(fname):

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension

    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit

    fname = fname + "x"

    return fname


# xlsx 파일로부터 전력사용량만 읽기
def merge_data(start, num):

    ## ismart 포맷에 맞춰서 전력수요값 뺴내기
    value_index = []
    for i in range(96):
        st = 'B' + str(i + 42)
        value_index.append(st)

    fbase = "C:\example\사용량종합현황"
    result = []

    # 엑셀파일들 호출
    for i in range(num):
        if i == 0:
            fname = fbase + ".xls"
        else:
            fname = fbase + " (" + str(i) + ").xls"

        fname = convert2xlsx(fname)

        wb = load_workbook(filename=fname, data_only=True)
        ws = wb.active

        day_demand = []

        for j in range(len(value_index)):
            #print(value_index[j])
            #print(ws[value_index[j]].value)
            day_demand.append(ws[(value_index[j])].value)

        result.append(day_demand)

    print("Data Merge 완료")

    ## 여기에 xlsx 파일로 저장하는거 추가하기 + 날짜 정보 추가해서 마무리



    print(result)

    pass

if __name__ == "__main__":
    merge_data(0, 7)


    # fname = "C:\Programming\RL\TetrisAI\사용량종합현황 (1).xls"
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # wb = excel.Workbooks.Open(fname)
    #
    # wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    #
    # wb.Close()  # FileFormat = 56 is for .xls extension
    # excel.Application.Quit



